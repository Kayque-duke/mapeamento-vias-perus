# ============================================================
# GENIUS EXCEL MASTER — Gerador de Planilha de Medições
# Nível: Excelência Imperial Absoluta
# ============================================================

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# 1. CRIAR WORKBOOK E CONFIGURAR
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Medicoes Kayque Rodrigues"

# ============================================================
# 2. SISTEMA DE DESIGN — PALETA PREMIUM DARK
# ============================================================
# Cores (sem #)
COR_BG_HEADER      = "0F1117"   # Header principal — preto profundo
COR_BG_SUBHEADER   = "1A1D27"   # Sub-header
COR_BG_LINHA_PAR   = "14141F"   # Linha par — escuro sutil
COR_BG_LINHA_IMPAR = "1E2230"   # Linha ímpar — contraste suave
COR_BG_STATS       = "0A0A0F"   # Área de estatísticas
COR_GOLD           = "C9A84C"   # Dourado principal — luxo
COR_GOLD_LIGHT     = "E5D49B"   # Dourado claro
COR_GOLD_DARK      = "8B7332"   # Dourado escuro
COR_TEXTO_PRIMARIO = "F5F5F7"   # Texto principal — branco suave
COR_TEXTO_SEC      = "A0A0B0"   # Texto secundário
COR_TEXTO_MUTED    = "6B6B80"   # Texto discreto
COR_AZUL_ACCENT    = "3B82F6"   # Azul acento para SI
COR_EMERALD        = "10B981"   # Verde esmeralda para conversões
COR_CYAN           = "22D3EE"   # Ciano para Imperial
COR_BORDER_SUTIL   = "2A2A3A"   # Borda sutil
COR_BORDER_GOLD    = "8B7332"   # Borda dourada
COR_VERMELHO       = "EF4444"   # Vermelho para alertas
COR_BG_TITULO      = "0A0A0F"   # Fundo do bloco de título

# Fills
fill_header      = PatternFill("solid", fgColor=COR_BG_HEADER)
fill_subheader   = PatternFill("solid", fgColor=COR_BG_SUBHEADER)
fill_par         = PatternFill("solid", fgColor=COR_BG_LINHA_PAR)
fill_impar       = PatternFill("solid", fgColor=COR_BG_LINHA_IMPAR)
fill_stats       = PatternFill("solid", fgColor=COR_BG_STATS)
fill_titulo      = PatternFill("solid", fgColor=COR_BG_TITULO)
fill_gold_accent = PatternFill("solid", fgColor="1F1A10")  # Fundo dourado sutil

# Fonts
font_titulo     = Font(name="Calibri", size=18, bold=True, color=COR_GOLD)
font_subtitulo  = Font(name="Calibri", size=11, bold=False, color=COR_TEXTO_SEC)
font_header     = Font(name="Calibri", size=11, bold=True, color=COR_GOLD)
font_subheader  = Font(name="Calibri", size=10, bold=True, color=COR_GOLD_LIGHT)
font_dados      = Font(name="Calibri", size=11, color=COR_TEXTO_PRIMARIO)
font_dados_num  = Font(name="Consolas", size=11, color=COR_TEXTO_PRIMARIO)
font_formula    = Font(name="Consolas", size=11, bold=True, color=COR_EMERALD)
font_conv       = Font(name="Consolas", size=11, bold=True, color=COR_CYAN)
font_unidade    = Font(name="Calibri", size=10, bold=True, color=COR_GOLD)
font_stats_label = Font(name="Calibri", size=10, bold=True, color=COR_GOLD_LIGHT)
font_stats_val  = Font(name="Consolas", size=11, bold=True, color=COR_EMERALD)
font_objeto_h   = Font(name="Calibri", size=11, bold=True, color=COR_AZUL_ACCENT)
font_rodape     = Font(name="Calibri", size=9, italic=True, color=COR_TEXTO_MUTED)

# Borders
borda_sutil = Border(
    left=Side(style="thin", color=COR_BORDER_SUTIL),
    right=Side(style="thin", color=COR_BORDER_SUTIL),
    top=Side(style="thin", color=COR_BORDER_SUTIL),
    bottom=Side(style="thin", color=COR_BORDER_SUTIL)
)
borda_header = Border(
    left=Side(style="thin", color=COR_BORDER_GOLD),
    right=Side(style="thin", color=COR_BORDER_GOLD),
    top=Side(style="medium", color=COR_GOLD),
    bottom=Side(style="medium", color=COR_GOLD)
)
borda_bottom_gold = Border(
    bottom=Side(style="medium", color=COR_GOLD)
)
borda_grupo_top = Border(
    top=Side(style="medium", color=COR_BORDER_GOLD),
    left=Side(style="thin", color=COR_BORDER_SUTIL),
    right=Side(style="thin", color=COR_BORDER_SUTIL),
    bottom=Side(style="thin", color=COR_BORDER_SUTIL)
)

# Alignment
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left   = Alignment(horizontal="left", vertical="center")
align_right  = Alignment(horizontal="right", vertical="center")

# ============================================================
# 3. CONFIGURAR LARGURAS DE COLUNA
# ============================================================
larguras = {
    "A": 22,   # Objeto
    "B": 16,   # Grandeza
    "C": 14,   # Leitura 1
    "D": 14,   # Leitura 2
    "E": 14,   # Leitura 3
    "F": 16,   # Média (SI)
    "G": 14,   # Unidade (SI)
    "H": 18,   # Conversão (Imperial)
    "I": 16,   # Unidade (Imperial)
}
for col, width in larguras.items():
    ws.column_dimensions[col].width = width

# ============================================================
# 4. BLOCO DE TÍTULO (Linhas 1-3)
# ============================================================
# Título principal
ws.merge_cells("A1:I1")
cell_titulo = ws["A1"]
cell_titulo.value = "RELATORIO DE MEDICOES — Kayque Rodrigues"
cell_titulo.font = font_titulo
cell_titulo.fill = fill_titulo
cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
cell_titulo.border = borda_bottom_gold
ws.row_dimensions[1].height = 42

# Aplicar fundo no título inteiro
for col_idx in range(1, 10):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = fill_titulo
    cell.border = borda_bottom_gold

# Subtítulo
ws.merge_cells("A2:I2")
cell_sub = ws["A2"]
cell_sub.value = "Sistema Internacional (SI) com conversao para Sistema Imperial | Tres leituras independentes por grandeza"
cell_sub.font = font_subtitulo
cell_sub.fill = fill_titulo
cell_sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 24

for col_idx in range(1, 10):
    cell = ws.cell(row=2, column=col_idx)
    cell.fill = fill_titulo

# Linha separadora vazia
ws.row_dimensions[3].height = 6
for col_idx in range(1, 10):
    cell = ws.cell(row=3, column=col_idx)
    cell.fill = fill_header

# ============================================================
# 5. CABEÇALHOS DA TABELA (Linha 4)
# ============================================================
cabecalhos = [
    "Objeto", "Grandeza",
    "Leitura 1", "Leitura 2", "Leitura 3",
    "Média (SI)", "Unidade (SI)",
    "Conversão (Imperial)", "Unidade (Imperial)"
]

HEADER_ROW = 4
ws.row_dimensions[HEADER_ROW].height = 36

for col_idx, header in enumerate(cabecalhos, 1):
    cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = borda_header

# ============================================================
# 6. DADOS — TODAS AS 13 LINHAS DE MEDIÇÃO
# ============================================================
# Formato: [Objeto, Grandeza, L1, L2, L3, fórmula_média, unid_SI, fórmula_conv, unid_Imperial]
# As fórmulas usam sintaxe EN (openpyxl traduz automaticamente)

dados = [
    # Linha 2 original → Linha 5 na planilha (offset +3)
    ["Ambiente de Medição", "Temperatura",
     22.5, 22.6, 22.5,
     '=ROUND(AVERAGE(C5:E5), 1)', "°C",
     '=ROUND((F5 * 1.8) + 32, 1)', "°F"],

    # iPhone
    ["iPhone", "Massa",
     240, 241, 240,
     '=ROUND(AVERAGE(C6:E6), 0)', "g",
     '=ROUND(F6 / 453.592, 4)', "lb"],

    ["iPhone", "Comprimento",
     16.00, 16.01, 16.00,
     '=ROUND(AVERAGE(C7:E7), 2)', "cm",
     '=ROUND(F7 / 2.54, 3)', "in"],

    ["iPhone", "Largura",
     7.76, 7.75, 7.76,
     '=ROUND(AVERAGE(C8:E8), 2)', "cm",
     '=ROUND(F8 / 2.54, 3)', "in"],

    ["iPhone", "Altura",
     0.78, 0.79, 0.78,
     '=ROUND(AVERAGE(C9:E9), 2)', "cm",
     '=ROUND(F9 / 2.54, 3)', "in"],

    # Livro Técnico
    ["Livro Técnico", "Massa",
     1250, 1251, 1250,
     '=ROUND(AVERAGE(C10:E10), 0)', "g",
     '=ROUND(F10 / 453.592, 4)', "lb"],

    ["Livro Técnico", "Comprimento",
     24.10, 24.11, 24.10,
     '=ROUND(AVERAGE(C11:E11), 2)', "cm",
     '=ROUND(F11 / 2.54, 3)', "in"],

    ["Livro Técnico", "Largura",
     17.20, 17.20, 17.19,
     '=ROUND(AVERAGE(C12:E12), 2)', "cm",
     '=ROUND(F12 / 2.54, 3)', "in"],

    ["Livro Técnico", "Altura",
     4.55, 4.56, 4.55,
     '=ROUND(AVERAGE(C13:E13), 2)', "cm",
     '=ROUND(F13 / 2.54, 3)', "in"],

    # Tubo Cilíndrico
    ["Tubo Cilíndrico", "Massa",
     205, 205, 206,
     '=ROUND(AVERAGE(C14:E14), 0)', "g",
     '=ROUND(F14 / 453.592, 4)', "lb"],

    ["Tubo Cilíndrico", "Comprimento",
     21.50, 21.52, 21.50,
     '=ROUND(AVERAGE(C15:E15), 2)', "cm",
     '=ROUND(F15 / 2.54, 3)', "in"],

    ["Tubo Cilíndrico", "Largura",
     7.60, 7.61, 7.60,
     '=ROUND(AVERAGE(C16:E16), 2)', "cm",
     '=ROUND(F16 / 2.54, 3)', "in"],

    ["Tubo Cilíndrico", "Altura",
     7.60, 7.60, 7.59,
     '=ROUND(AVERAGE(C17:E17), 2)', "cm",
     '=ROUND(F17 / 2.54, 3)', "in"],
]

# Definir quais linhas são "primeira do grupo" (para borda de separação)
grupos_inicio = {
    0: "Ambiente de Medição",
    1: "iPhone",
    5: "Livro Técnico",
    9: "Tubo Cilíndrico"
}

DATA_START_ROW = 5  # Linha onde começam os dados

for idx, row_data in enumerate(dados):
    row_num = DATA_START_ROW + idx
    ws.row_dimensions[row_num].height = 28

    # Determinar fill (zebra striping por grupo)
    is_par = (idx % 2 == 0)
    fill_row = fill_par if is_par else fill_impar

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill = fill_row
        cell.alignment = align_center

        # Determinar borda
        if idx in grupos_inicio:
            cell.border = borda_grupo_top
        else:
            cell.border = borda_sutil

        # === FONTES POR COLUNA ===
        if col_idx == 1:  # Objeto
            cell.font = font_objeto_h
            cell.alignment = align_left
        elif col_idx == 2:  # Grandeza
            cell.font = font_dados
        elif col_idx in (3, 4, 5):  # Leituras
            cell.font = font_dados_num
            # Formatação numérica condicional
            if row_data[1] == "Massa":
                cell.number_format = '#,##0'
            elif row_data[1] == "Temperatura":
                cell.number_format = '0.0'
            else:
                cell.number_format = '0.00'
        elif col_idx == 6:  # Média (SI) — fórmula
            cell.font = font_formula
            if row_data[1] == "Massa":
                cell.number_format = '#,##0'
            elif row_data[1] == "Temperatura":
                cell.number_format = '0.0'
            else:
                cell.number_format = '0.00'
        elif col_idx == 7:  # Unidade SI
            cell.font = font_unidade
        elif col_idx == 8:  # Conversão Imperial — fórmula
            cell.font = font_conv
            if row_data[1] == "Massa":
                cell.number_format = '0.0000'
            elif row_data[1] == "Temperatura":
                cell.number_format = '0.0'
            else:
                cell.number_format = '0.000'
        elif col_idx == 9:  # Unidade Imperial
            cell.font = font_unidade

# ============================================================
# 7. BLOCO DE ESTATÍSTICAS (Abaixo dos dados)
# ============================================================
STATS_START = DATA_START_ROW + len(dados) + 1  # Pula uma linha

# Linha separadora
ws.row_dimensions[STATS_START - 1].height = 6
for col_idx in range(1, 10):
    cell = ws.cell(row=STATS_START - 1, column=col_idx)
    cell.fill = fill_header

# Título das estatísticas
ws.merge_cells(f"A{STATS_START}:I{STATS_START}")
cell_stat_title = ws.cell(row=STATS_START, column=1, value="📈 ESTATÍSTICAS RESUMIDAS")
cell_stat_title.font = Font(name="Calibri", size=14, bold=True, color=COR_GOLD)
cell_stat_title.fill = fill_stats
cell_stat_title.alignment = Alignment(horizontal="center", vertical="center")
cell_stat_title.border = borda_bottom_gold
ws.row_dimensions[STATS_START].height = 34

for col_idx in range(1, 10):
    ws.cell(row=STATS_START, column=col_idx).fill = fill_stats
    ws.cell(row=STATS_START, column=col_idx).border = borda_bottom_gold

# Estatísticas por objeto
stats_data = [
    ["iPhone — Massa Média", f'=F6', "g", f'=H6', "lb"],
    ["iPhone — Volume Aprox. (C×L×A)", f'=ROUND(F7*F8*F9, 2)', "cm³", f'=ROUND(H7*H8*H9, 4)', "in³"],
    ["Livro Técnico — Massa Média", f'=F10', "g", f'=H10', "lb"],
    ["Livro Técnico — Volume Aprox. (C×L×A)", f'=ROUND(F11*F12*F13, 2)', "cm³", f'=ROUND(H11*H12*H13, 4)', "in³"],
    ["Tubo Cilíndrico — Massa Média", f'=F14', "g", f'=H14', "lb"],
    ["Tubo Cilíndrico — Volume Aprox. (π×(D/2)²×C)", f'=ROUND(PI()*(F16/2)^2*F15, 2)', "cm³", f'=ROUND(PI()*(H16/2)^2*H15, 4)', "in³"],
    ["Total de Medições Realizadas", f'=COUNTA(C5:E17)', "leituras", "", ""],
    ["Temperatura do Ensaio", f'=F5', "°C", f'=H5', "°F"],
]

for i, stat in enumerate(stats_data):
    row_num = STATS_START + 1 + i
    ws.row_dimensions[row_num].height = 26
    is_par = (i % 2 == 0)
    fill_row = fill_par if is_par else fill_impar

    # Label (merge A:C)
    ws.merge_cells(f"A{row_num}:C{row_num}")
    cell_label = ws.cell(row=row_num, column=1, value=stat[0])
    cell_label.font = font_stats_label
    cell_label.fill = fill_row
    cell_label.alignment = Alignment(horizontal="right", vertical="center")
    cell_label.border = borda_sutil
    for merge_col in range(2, 4):
        ws.cell(row=row_num, column=merge_col).fill = fill_row
        ws.cell(row=row_num, column=merge_col).border = borda_sutil

    # Valor SI (merge D:E)
    ws.merge_cells(f"D{row_num}:E{row_num}")
    cell_val_si = ws.cell(row=row_num, column=4, value=stat[1])
    cell_val_si.font = font_stats_val
    cell_val_si.fill = fill_row
    cell_val_si.alignment = align_center
    cell_val_si.border = borda_sutil
    if "Volume" in stat[0]:
        cell_val_si.number_format = '#,##0.00'
    elif "Medições" in stat[0]:
        cell_val_si.number_format = '0'
    ws.cell(row=row_num, column=5).fill = fill_row
    ws.cell(row=row_num, column=5).border = borda_sutil

    # Unidade SI
    cell_un_si = ws.cell(row=row_num, column=6, value=stat[2])
    cell_un_si.font = font_unidade
    cell_un_si.fill = fill_row
    cell_un_si.alignment = align_center
    cell_un_si.border = borda_sutil

    # Valor Imperial (merge G:H)
    ws.merge_cells(f"G{row_num}:H{row_num}")
    cell_val_imp = ws.cell(row=row_num, column=7, value=stat[3])
    cell_val_imp.font = Font(name="Consolas", size=11, bold=True, color=COR_CYAN)
    cell_val_imp.fill = fill_row
    cell_val_imp.alignment = align_center
    cell_val_imp.border = borda_sutil
    if "Volume" in stat[0]:
        cell_val_imp.number_format = '0.0000'
    ws.cell(row=row_num, column=8).fill = fill_row
    ws.cell(row=row_num, column=8).border = borda_sutil

    # Unidade Imperial
    cell_un_imp = ws.cell(row=row_num, column=9, value=stat[4])
    cell_un_imp.font = font_unidade
    cell_un_imp.fill = fill_row
    cell_un_imp.alignment = align_center
    cell_un_imp.border = borda_sutil

# ============================================================
# 8. RODAPÉ (removido conforme solicitado)
# ============================================================

# ============================================================
# 9. CONFIGURAÇÕES DA PLANILHA
# ============================================================
# Congelar painéis (abaixo do header)
ws.freeze_panes = f"A{DATA_START_ROW}"

# Configurar tab color (dourado)
ws.sheet_properties.tabColor = COR_GOLD

# Configurar visualização (zoom)
ws.sheet_view.zoomScale = 110

# Remover gridlines
ws.sheet_view.showGridLines = False

# Configuração de impressão
ws.print_options.horizontalCentered = True
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1

# ============================================================
# 10. SALVAR
# ============================================================
output_path = r"c:\Users\gabri\Nova pasta (3)\Medicoes_Kayque_Rodrigues.xlsx"
wb.save(output_path)
print("Planilha gerada com sucesso!")
print(f"Caminho: {output_path}")
print(f"{len(dados)} linhas de medicoes + {len(stats_data)} estatisticas")
print("Design: Dark Mode Premium com paleta dourada")
print("Formulas: ROUND(AVERAGE()) para medias + conversoes SI para Imperial")
